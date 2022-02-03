from openpyxl import load_workbook, Workbook


# Get character categories from Veale
def get_veales():
    wb = load_workbook(filename='../data/Veale_db/Veale\'s category actions.xlsx')
    work_sheet = wb.active
    characters = [work_sheet[i][1].value for i in range(2, 451)]        # start from 2, don't include header row
    return work_sheet, characters


# Process fairytale dataset into one string
def process_fairy_tales():
    fairy_tales = ''
    with open('../data/Fairytales_db/merged_clean.txt') as file:
        lines = file.readlines()
        lines = [line.strip() for line in lines]
        for line in lines:
            fairy_tales += line.lower()
    return fairy_tales


# For each character, count how many times it occurs in fairytale dataset and put >1 into txt file
def find_fairy_tale_characters(fairy_tales, characters):
    # adds spaces for 'word bounds'
    counted_characters = [[character, fairy_tales.count(' ' + character.lower() + ' ')] for character in characters if fairy_tales.count(' ' + character.lower() + ' ') > 0]
    counted_characters = sorted(counted_characters, key=lambda x: x[1], reverse=True)

    # Write all left over characters to file
    f = open("../data/Own_db/characters.txt", "w")
    for i in range(0, len(counted_characters)):
        f.write(str(i + 1) + '. ' + counted_characters[i][0] + ' : ' + str(counted_characters[i][1]) + ' times\n')
    f.close()
    return counted_characters


def extract_to_xlsx(character_names, file_name, work_sheet):
    # create new Excel file
    new_wb = Workbook()
    new_ws = new_wb.active
    # dict for knowing the row in which character is
    char_rows = {work_sheet[i][1].value: i for i in range(1, 451) if work_sheet[i][1].value in character_names}

    # copy header from old workbook to new workbook
    new_ws.cell(row=1, column=1, value=work_sheet[1][0].value)
    new_ws.cell(row=1, column=2, value=work_sheet[1][1].value)
    new_ws.cell(row=1, column=3, value=work_sheet[1][2].value)
    new_ws.cell(row=1, column=4, value=work_sheet[1][3].value)

    for i in range(0, len(character_names)):
        char = character_names[i]
        old_index = char_rows[char]
        new_index = i + 2

        # copy row from old workbook to new workbook
        new_ws.cell(row=new_index, column=1, value=work_sheet[old_index][0].value)
        new_ws.cell(row=new_index, column=2, value=work_sheet[old_index][1].value)
        new_ws.cell(row=new_index, column=3, value=work_sheet[old_index][2].value)
        new_ws.cell(row=new_index, column=4, value=work_sheet[old_index][3].value)

    new_wb.save(filename='../data/Own_db/' + file_name)


def get_fairytale_characters(work_sheet, characters):
    _fairy_tales = process_fairy_tales()
    _counted_characters = find_fairy_tale_characters(_fairy_tales, characters)
    _character_list = [item[0] for item in _counted_characters]
    extract_to_xlsx(_character_list, 'extracted_category_actions.xlsx', work_sheet)


def get_humanlike_characters(work_sheet):
    with open('../data/Own_db/humanlike_characters.txt') as f:
        characters = f.readlines()
        characters = [item.strip() for item in characters]
        extract_to_xlsx(characters, 'humanlike_extracted_category_actions.xlsx', work_sheet)


veales_worksheet, veales_characters = get_veales()
# get_fairytale_characters(veales_worksheet, veales_characters)
get_humanlike_characters(veales_worksheet)
