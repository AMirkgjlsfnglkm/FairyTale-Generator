import random as rd
from openpyxl import load_workbook

vowels = ['a', 'e', 'i', 'u', 'o']
names = ["Sam", "Alex", "Charlie", "Chris", "Frankie", "Jackie", "Kendall"]


class Character:
    name = ""
    char_type = ""
    trait = ""
    article = ""

    def __init__(self, char_type, trait, char, other_name=""):
        self.create_name(other_name)
        self.char_type = char_type
        self.trait = trait
        self.determine_article(char)

    def create_name(self, other_name):
        if other_name == "":
            name = rd.choice(names)
        else:
            copy_of_names = names.copy()
            copy_of_names.remove(other_name)
            name = rd.choice(copy_of_names)
        self.name = name

    def determine_article(self, char):
        if char == "a":
            article = "a"
            if self.trait[0] in vowels:
                article = "an"
            self.article = article
        else:
            article = "a"
            if self.char_type[0] in vowels:
                article = "an"
            self.article = article

    def name_call(self, is_object=True, choice=False):
        if choice and is_object:
            return "they"
        elif choice and not is_object:
            return "them"
        return "the " + self.char_type


# The story class holds all function needed to create a story,
# The seed given at creation will be used in the random elements of the story
class Story:

    # These variables store the different story elements
    intro = ""
    first_midpoint = []
    second_midpoint = []
    character_a = None
    character_b = None
    need = ""
    ending = ""
    template_start = ""
    template_intro_a = ""
    template_intro_b = ""
    template_need = ""          # not used yet
    template_end = ""

    def __init__(self, seed):
        rd.seed(seed)

        self.init_templates()

        self.create_second_midpoint()
        self.create_first_midpoint()

        self.create_character_a()
        self.create_character_b()
        self.create_need()
        self.create_ending()

        self.return_story()

    # This function creates the first midpoint of the story: a truple of verbs
    def create_second_midpoint(self):
        # Open the "pruned script midpoints" sheet
        workbook = load_workbook(filename="../data/Own_db/Our midpoints pruned 2.xlsx")
        work_sheet = workbook.active

        # Choose a random row in the worksheet
        random_index = rd.randint(1, 124)

        # Store the first verb in the row
        midpoint = [work_sheet[random_index][0].value]

        # Choose a random one of the second verbs in the row
        verb2 = work_sheet[random_index][1].value.split(",")
        midpoint.append(rd.choice(verb2))

        # Choose a random one of the third verbs in the row
        verb3 = work_sheet[random_index][2].value.split(",")
        midpoint.append(rd.choice(verb3))

        # Store the second midpoint
        self.second_midpoint = midpoint

    # This function creates the first midpoint of the story, a truple of verbs, based upon the second midpoint
    def create_first_midpoint(self):
        # Open the "Script midpoints" sheet
        workbook = load_workbook(filename="../data/Veale_db/Veale_s script midpoints.xlsx")
        work_sheet = workbook.active

        # Retrieve the last verb of the first midpoint
        ending_verb = self.second_midpoint[0]

        # Find all the options for rows ending with the verb
        options = []
        for row in range(2, 2761):
            if ending_verb in work_sheet[row][2].value.split(", "):
                options.append(row)

        # If there are options, find first midpoint
        midpoint_found = False
        while len(options) != 0 and not midpoint_found:
            # Choose a random one of these rows
            row = rd.choice(options)

            # Store the first verb in the row if it's not a duplicate of the second midpoint
            verb1 = work_sheet[row][0].value.split(", ")
            verb1 = [verb for verb in verb1 if verb not in self.second_midpoint]
            if len(verb1) == 0:
                continue
            midpoint = [rd.choice(verb1)]

            # Choose a random one of the second verbs in the row if it's not a duplicate of the second midpoint
            verb2 = work_sheet[row][1].value.split(", ")
            verb2 = [verb for verb in verb2 if verb not in self.second_midpoint]
            if len(verb2) == 0:
                continue
            midpoint.append(rd.choice(verb2))

            # Choose the first verb in the second midpoint as third verb
            verb3 = self.second_midpoint[0]
            midpoint.append(verb3)

            # If this point is reached, all midpoint verbs have been selected that are not duplicates
            midpoint_found = True
            self.first_midpoint = midpoint

        if not midpoint_found:
            print("No first midpoint: starting over")

    # This function creates a character A and a trait based upon the first midpoint, as a subject
    def create_character_a(self):
        # Open the "Category actions" sheet
        workbook = load_workbook(filename="../data/Own_db/humanlike_extracted_category_actions.xlsx")
        work_sheet = workbook.active

        # Retrieve the midpoint verbs from the first midpoint
        midpoint_options = self.first_midpoint.copy()
        options = []

        # Until a character is found or until we are out of verbs
        while len(options) < 1 and len(midpoint_options) > 0:
            # Pick one of the remaining verbs
            verb = rd.choice(midpoint_options)

            # Remove it from the possible options
            midpoint_options.remove(verb)

            # Find the rows that contain the verb
            options = []
            for row in range(1, 150):
                if work_sheet[row][2].value is not None:
                    if verb in work_sheet[row][2].value.split(", "):
                        options.append(row)

        # If no possible options were found
        if len(options) < 1:
            print("No character found: taking a default one")
            character = "prince"
        else:
            # Choose a random one of the possible character options
            character = work_sheet[rd.choice(options)][1].value.lower()

        # Open the "Quality inventory" sheet
        workbook = load_workbook(filename="../data/Veale_db/Veale_s quality inventory.xlsx")
        work_sheet = workbook.active

        # Retrieve the midpoint verbs from the first midpoint
        midpoint_options = self.first_midpoint.copy()
        options = []

        # Until a trait is found or until we are out of verbs
        while len(options) < 1 and len(midpoint_options) > 0:
            # Pick one of the remaining verbs
            verb = rd.choice(midpoint_options)

            # Remove it from the possible options
            midpoint_options.remove(verb)

            # Find the rows that contain the verb
            options = []
            for row in range(1, 1973):
                if work_sheet[row][2].value is not None:
                    if verb in work_sheet[row][2].value.split(", "):
                        options.append(row)

        # If no possible options were found
        if len(options) < 1:
            print("No trait found: taking a default one")
            trait = "brave"
        else:
            # Choose a random one of the possible trait options
            trait = work_sheet[rd.choice(options)][0].value

        # Store the character, trait and article
        obj_character = Character(character, trait, "a")
        self.character_a = obj_character

    # This function creates a character B and a trait based upon the first midpoint, as an object
    def create_character_b(self):
        # Open the "Category actions" sheet
        workbook = load_workbook(filename="../data/Own_db/humanlike_extracted_category_actions.xlsx")
        work_sheet = workbook.active

        # Retrieve the midpoint verbs from the first midpoint
        midpoint_options = self.first_midpoint.copy()
        options = []

        # Until a character is found or until we are out of verbs
        while len(options) < 1 and len(midpoint_options) > 0:
            # Pick one of the remaining verbs
            verb = rd.choice(midpoint_options)

            # Remove it from the possible options
            midpoint_options.remove(verb)

            # Find the rows that contain the verb
            options = []
            for row in range(1, 150):
                if work_sheet[row][3].value is not None:
                    if verb in work_sheet[row][3].value.split(", "):
                        options.append(row)

        # If no possible options were found
        if len(options) < 1:
            print("No character found: taking a default one")
            character = "princess"
        else:
            # Choose a random one of the possible character options
            character = work_sheet[rd.choice(options)][1].value.lower()

        # Open the "Quality inventory" sheet
        workbook = load_workbook(filename="../data/Veale_db/Veale_s quality inventory.xlsx")
        work_sheet = workbook.active

        # Retrieve the midpoint verbs from the first midpoint
        midpoint_options = self.first_midpoint.copy()
        options = []

        # Until a trait is found or until we are out of verbs
        while len(options) < 1 and len(midpoint_options) > 0:
            # Pick one of the remaining verbs
            verb = rd.choice(midpoint_options)

            # Remove it from the possible options
            midpoint_options.remove(verb)

            # Find the rows that contain the verb
            options = []
            for row in range(1, 1973):
                if work_sheet[row][3].value is not None:
                    if verb in work_sheet[row][3].value.split(", "):
                        options.append(row)

        # If no possible options were found
        if len(options) < 1:
            print("No trait found: taking a default one")
            trait = "smart"
        else:
            # Choose a random one of the possible trait options
            trait = work_sheet[rd.choice(options)][0].value

        obj_character = Character(character, trait, "b", self.character_a.name)
        self.character_b = obj_character

    # This function creates a need for character A based on the midpoints.
    def create_need(self):
        workbook = load_workbook(filename="../data/Own_db/Our needs.xlsx")
        work_sheet = workbook.active

        verb = self.second_midpoint[2]
        need = ""
        for row in range(2, 12):
            if work_sheet[row][0].value == verb:
                need = rd.choice(work_sheet[row][1].value.split(", "))
                if 'C' in need:
                    need = need.replace("C", self.template_need)
                else:
                    need = 'A ' + need

        if need == "":
            print("No need found: taking standard one")
            need = "A wanted to change their boring ways"
        self.need = need

    # This function generates an ending based on the second midpoint
    def create_ending(self):
        # Open the "Closing bookend actions" sheet
        workbook = load_workbook(filename="../data/Veale_db/Veale_s closing bookend actions.xlsx")
        work_sheet = workbook.active

        # Retrieve the last verb of the second midpoint
        verb = self.second_midpoint[2]

        # Find the row in which this verb resides and pick one of the options of closing actions.
        for row in range(1, 244):
            if work_sheet[row][0].value == verb:
                self.ending = rd.choice(work_sheet[row][2].value.split(", "))

        # If no row was found using the verb
        if self.ending == "":
            print("No bookend found: taking standard one")
            self.ending = self.template_end

    # This function chooses templates for the story at random
    def init_templates(self):
        workbook = load_workbook(filename='../data/Own_db/Templates.xlsx')
        worksheet = workbook.active

        # Make lists with all the possible template sentences for each topic/sentence required
        start_sentence = [worksheet[i][0].value for i in range(2, 8)]
        intro_character_a = [worksheet[i][1].value for i in range(2, 4)]
        intro_character_b = [worksheet[i][2].value for i in range(2, 8)]
        need = [worksheet[i][3].value for i in range(2, 8)]
        ending = [worksheet[i][4].value for i in range(2, 6)]

        # Randomly select template sentences
        self.template_start = rd.choice(start_sentence)
        self.template_intro_a = rd.choice(intro_character_a)
        self.template_intro_b = rd.choice(intro_character_b)
        self.template_need = rd.choice(need)
        self.template_end = rd.choice(ending)

    # This function prints the story in a readable format.
    def return_story(self):
        character_a = self.character_a
        character_b = self.character_b

        # Start sentence
        print(self.template_start, self.template_intro_a, character_a.article, character_a.trait, character_a.char_type,
              "called", character_a.name + ".")

        if "A" in self.need and "B" in self.need:
            # Introduction of character B
            print(self.template_intro_b, character_b.article, character_b.char_type,
                  "that was", character_b.trait + ".")
            # Introduction of need
            need_intro = (self.need.replace("A", character_a.name_call())
                          .replace("B", character_b.name_call()))
            need_intro = need_intro[0].capitalize() + need_intro[1:]
            print(need_intro + ".")
        else:
            # Introduction of need
            need_intro = (self.need.replace("A", character_a.name_call())
                          .replace("B", character_b.name_call()))
            need_intro = need_intro[0].capitalize() + need_intro[1:]
            print(need_intro + ".")
            # Introduction of character B
            print(self.template_intro_b, character_b.article, character_b.char_type,
                  "that was", character_b.trait + ".")

        # First midpoint idiomatized
        for verb in self.first_midpoint:
            verb_sentence = self.util_idiomatize(verb)
            if "A" in verb_sentence and "B" in verb_sentence:
                a_object = verb_sentence.index("A") < verb_sentence.index("B")
                b_object = not a_object
            else:
                a_object = "A" in verb_sentence
                b_object = not a_object
            choice = rd.choice([(False, False), (True, False), (False, True)])
            verb_replaced = verb_sentence.replace("A", character_a.name_call(a_object, choice[0]))\
                .replace("B", character_b.name_call(b_object, choice[1]))
            verb_replaced = verb_replaced[0].capitalize() + verb_replaced[1:]
            print(verb_replaced)

        # Second midpoint idiomatized
        for verb in self.second_midpoint[1:]:
            verb_sentence = self.util_idiomatize(verb)
            if "A" in verb_sentence and "B" in verb_sentence:
                a_object = verb_sentence.index("A") < verb_sentence.index("B")
                b_object = not a_object
            else:
                a_object = "A" in verb_sentence
                b_object = not a_object
            choice = rd.choice([(False, False), (True, False), (False, True)])
            verb_replaced = verb_sentence.replace("A", character_a.name_call(a_object, choice[0]))\
                .replace("B", character_b.name_call(b_object, choice[1]))
            verb_replaced = verb_replaced[0].capitalize() + verb_replaced[1:]
            print(verb_replaced)

        # Ending
        ending_sentence = self.ending.replace("A", character_a.name_call())\
            .replace("B", character_b.name_call())
        ending_sentence = ending_sentence[0].capitalize() + ending_sentence[1:]
        print(ending_sentence + ".")

    # This function returns the idiomatized version of the input verb
    @staticmethod
    def util_idiomatize(verb):
        # Open the "Idiomatic actions" sheet
        workbook = load_workbook(filename="../data/Veale_db/Veale_s idiomatic actions.xlsx")
        work_sheet = workbook.active

        # If the verb is reversed
        if verb[0] == "*":
            verb = verb[1:]
            # Find the row with the verb and return a random idiomatized version from it
            for row in range(2, 820):
                if work_sheet[row][0].value == verb:
                    # Reverse the characters
                    return rd.choice(work_sheet[row][4].value.split(", "))\
                        .replace("A", "C")\
                        .replace("B", "A")\
                        .replace("C", "B") + "."
        else:
            # Find the row with the verb and return a random idiomatized version from it
            for row in range(2, 820):
                if work_sheet[row][0].value == verb:
                    return rd.choice(work_sheet[row][4].value.split(", ")) + "."

        # If none are found
        return verb + "."


# for i in range(102, 120):
#     print(i)
#     Story(i)
#     print('\n', '\n')
new_seed = rd.randint(0, 1000000)
print('A story will be generated with seed:', new_seed)
Story(new_seed)